import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# 1 - Entrar na planilha e extrair INFORMAÇÕES do cliente

planinha_clientes = openpyxl.load_workbook('') #Abre a planilha
pagina_clientes = planinha_clientes[''] #PAGINA DA PLANILHA

driver = webdriver.Chrome()
driver.get('') #Link do site

for linha in pagina_clientes.iter_rows(min_row=2,values_only=True):
    nome, valor, cpf, vencimento = linha
    

    # 2 - Entrar no site e usar o CPF da planilha para pesquisar o status de pagamento do cliente

    
    sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH, ) #"//tag[@atributo='valor']"
    sleep(1)
    campo_pesquisa.send_keys(cpf)
    sleep(1)

    # 3 - Verificar Status 
    botao_pesquisar = driver.find_element(By.XPATH, ) #"//tag[@atributo='valor']" CLICK NO BOTAO DE CONSULTA
    sleep(1)
    botao_pesquisar.click()
    sleep(3)
    status = driver.find_element(By.XPATH, ) #"//tag[@atributo='valor']" LÊ STATUS DA CONSULTA
    if status.text == '': 
        # 4 - Se Status "Positivo",pegar DATA DO PAGAMENTO e  MÉTODO DO PAGAMENTO
        data_pagamento = driver.find_element(By.XPATH, ) #"//tag[@atributo='valor'] LÊ DATA DO PAGAMENTO
        metodo_pagamento = driver.find_element(By.XPATH, ) #"//tag[@atributo='valor'] LÊ METODO DO PAGAMENTO

        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('') #Abre a planilha
        pagina_fechamento = planilha_fechamento[''] #Página da planilha 

        pagina_fechamento.append([nome, valor, cpf, vencimento,'em dia',
        metodo_pagamento_limpo, data_pagamento_limpo])

        planilha_fechamento.save()

    else:
        # 5 - Caso Status Contrário, Colocar como pendente
        planilha_fechamento = openpyxl.load_workbook('')  #Abre a planilha
        pagina_fechamento = planilha_fechamento[''] #Página da planilha
        
        pagina_fechamento.append([nome, valor, cpf, vencimento,'pendente'])
        planilha_fechamento.save()
